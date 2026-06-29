#!/usr/bin/env python3
"""
Step 3 — build the deduplicated, categorized keyword XLSX (and CSV mirror).

Opportunity-Score Formula
-------------------------
::

    base = volume / (1 + best_competitor_position)

Then a positional penalty multiplier is applied based on where the SITE
currently ranks for the keyword:

    - our_position in 1-10   -> base * 0.30   (already winning; deprioritise)
    - our_position in 11-50  -> base * 0.70   (good push opportunity)
    - our_position > 50      -> base * 1.00   (full weight)
    - our_position is None   -> base * 1.00   (full weight; pure gap)

This biases the list toward gaps and near-misses — exactly the keywords where
incremental work moves rankings the fastest.

Sheets
------
- **High Opportunity** — top 25% by opportunity_score, NOT in our top 10,
  search_volume > 100. Sorted by opportunity_score desc.
- **Hidden Gems** — search_volume in [50, 500], exactly 1 competitor ranks,
  KD <= 30 if known. Sorted by opportunity_score desc.
- **High Volume** — top 100 by raw search_volume. Sorted by search_volume desc.
- **All Keywords** — full dedup'd list. Sorted by opportunity_score desc.

Inputs
------
- ``--vault``: vault root (reads from ``.raw/sources/dataforseo/``).
- ``--out-xlsx`` / ``--out-csv``: explicit paths (defaults derived from vault).

Outputs
-------
- ``<vault>/keywords-<date>.xlsx`` (4 sheets) — vault root, user-facing
- ``<vault>/keywords-<date>.csv`` (All Keywords mirror) — vault root, user-facing
- Console summary: row counts per sheet, top 5 from High Opportunity.

Dependency: ``openpyxl``.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    sys.stderr.write(
        "ERROR: openpyxl not installed. Install it with:\n    pip install openpyxl\n"
    )
    sys.exit(2)


def _normalize_keyword(kw: str) -> str:
    return re.sub(r"\s+", " ", kw.strip().lower())


def _opportunity_score(volume: int, best_pos: int | None, our_pos: int | None) -> float:
    """See module docstring for the formula."""
    if best_pos is None or volume <= 0:
        return 0.0
    base = volume / (1 + best_pos)
    if our_pos is None:
        penalty = 1.0
    elif our_pos <= 10:
        penalty = 0.3
    elif our_pos <= 50:
        penalty = 0.7
    else:
        penalty = 1.0
    return round(base * penalty, 2)


def _load_competitor_files(raw_dir: Path) -> list[Path]:
    return sorted(raw_dir.glob("competitor-kw-*.json"))


def _load_site_file(raw_dir: Path) -> Path | None:
    candidates = sorted(raw_dir.glob("site-ranked-keywords-*.json"))
    return candidates[-1] if candidates else None


def _iter_items(consolidated: dict[str, Any]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for page in consolidated.get("pages", []) or []:
        for item in page.get("items", []) or []:
            out.append(item)
    return out


def _extract_row(domain: str, item: dict[str, Any]) -> dict[str, Any] | None:
    kw_data = item.get("keyword_data") or {}
    kw_info = kw_data.get("keyword_info") or {}
    props = kw_data.get("keyword_properties") or {}
    intent_info = kw_data.get("search_intent_info") or {}
    serp_info = kw_data.get("serp_info") or {}
    ranked = item.get("ranked_serp_element") or {}
    serp_item = ranked.get("serp_item") or {}
    keyword = kw_data.get("keyword") or ""
    if not keyword:
        return None
    return {
        "domain": domain,
        "keyword": keyword,
        "keyword_norm": _normalize_keyword(keyword),
        "rank_group": serp_item.get("rank_group"),
        "ranking_url": serp_item.get("url"),
        "search_volume": kw_info.get("search_volume") or 0,
        "cpc": kw_info.get("cpc"),
        "competition": kw_info.get("competition"),
        "competition_level": kw_info.get("competition_level"),
        "kd": props.get("keyword_difficulty"),
        "intent": intent_info.get("main_intent"),
        "serp_features": ", ".join(serp_info.get("serp_item_types") or []),
    }


def _aggregate(
    site_root: str | None,
    site_rows: list[dict[str, Any]],
    competitor_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Group by normalized keyword, compute per-row analytics."""
    site_by_kw: dict[str, dict[str, Any]] = {}
    for r in site_rows:
        prior = site_by_kw.get(r["keyword_norm"])
        if not prior or (r.get("rank_group") or 999) < (prior.get("rank_group") or 999):
            site_by_kw[r["keyword_norm"]] = r

    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for r in competitor_rows:
        grouped[r["keyword_norm"]].append(r)

    out: list[dict[str, Any]] = []
    for norm, rows in grouped.items():
        # Pick a "primary" row for keyword metadata — the highest-volume one.
        primary = max(rows, key=lambda r: int(r.get("search_volume") or 0))
        comp_domains = sorted({r["domain"] for r in rows})
        # best competitor position
        best = min(rows, key=lambda r: int(r.get("rank_group") or 999))
        best_pos = int(best.get("rank_group") or 999) if best.get("rank_group") else None
        site_row = site_by_kw.get(norm)
        our_pos = int(site_row.get("rank_group")) if site_row and site_row.get("rank_group") else None
        volume = int(primary.get("search_volume") or 0)
        score = _opportunity_score(volume, best_pos, our_pos)
        out.append({
            "keyword": primary["keyword"],
            "search_volume": volume,
            "kd": primary.get("kd"),
            "cpc": primary.get("cpc"),
            "competition": primary.get("competition"),
            "intent": primary.get("intent"),
            "serp_features": primary.get("serp_features"),
            "competitor_count": len(comp_domains),
            "best_competitor_position": best_pos,
            "best_competitor_domain": best.get("domain"),
            "best_competitor_url": best.get("ranking_url"),
            "all_competitors": "; ".join(f"{r['domain']} #{r.get('rank_group')}" for r in sorted(rows, key=lambda x: int(x.get("rank_group") or 999))),
            "our_position": our_pos,
            "our_url": site_row.get("ranking_url") if site_row else None,
            "opportunity_score": score,
        })
    out.sort(key=lambda r: (-r["opportunity_score"], -r["search_volume"]))
    return out


def _categorize(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    if not rows:
        return {"High Opportunity": [], "Hidden Gems": [], "High Volume": [], "All Keywords": []}
    scores_sorted = sorted([r["opportunity_score"] for r in rows], reverse=True)
    quartile_idx = max(1, len(scores_sorted) // 4) - 1
    threshold = scores_sorted[quartile_idx]

    high_opp = [r for r in rows
                if r["opportunity_score"] >= threshold
                and (r["our_position"] is None or r["our_position"] > 10)
                and r["search_volume"] > 100]
    high_opp.sort(key=lambda r: -r["opportunity_score"])

    hidden_gems = [r for r in rows
                   if 50 <= r["search_volume"] <= 500
                   and r["competitor_count"] == 1
                   and (r["kd"] is None or (isinstance(r["kd"], (int, float)) and r["kd"] <= 30))]
    hidden_gems.sort(key=lambda r: -r["opportunity_score"])

    high_volume = sorted(rows, key=lambda r: -r["search_volume"])[:100]

    return {
        "High Opportunity": high_opp,
        "Hidden Gems": hidden_gems,
        "High Volume": high_volume,
        "All Keywords": rows,
    }


# --- Output writers ---------------------------------------------------------
_HEADERS = [
    "keyword", "search_volume", "kd", "cpc", "competition", "intent",
    "competitor_count", "best_competitor_position", "best_competitor_domain",
    "best_competitor_url", "all_competitors", "our_position", "our_url",
    "opportunity_score", "serp_features",
]


def _write_xlsx(path: Path, sheets: dict[str, list[dict[str, Any]]]) -> None:
    wb = Workbook()
    # Remove default sheet then add ours in order.
    default = wb.active
    wb.remove(default)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1B5273")
    header_align = Alignment(horizontal="left", vertical="center")

    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        # write header
        for col, h in enumerate(_HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        # rows
        for r_idx, row in enumerate(rows, start=2):
            for c_idx, h in enumerate(_HEADERS, 1):
                ws.cell(row=r_idx, column=c_idx, value=row.get(h))
        # cosmetics
        ws.freeze_panes = "A2"
        if rows:
            ws.auto_filter.ref = ws.dimensions
        # column widths (cheap)
        for c_idx, h in enumerate(_HEADERS, 1):
            max_len = len(h)
            for r_idx in range(2, min(2 + 50, ws.max_row + 1)):
                v = ws.cell(row=r_idx, column=c_idx).value
                if v is not None:
                    max_len = max(max_len, min(len(str(v)), 60))
            ws.column_dimensions[get_column_letter(c_idx)].width = max_len + 2
        # color scale on opportunity_score (column index from _HEADERS)
        score_col = _HEADERS.index("opportunity_score") + 1
        score_letter = get_column_letter(score_col)
        if rows:
            ws.conditional_formatting.add(
                f"{score_letter}2:{score_letter}{len(rows) + 1}",
                ColorScaleRule(
                    start_type="min", start_color="FFFFFF",
                    mid_type="percentile", mid_value=50, mid_color="FFCFA8",
                    end_type="max", end_color="C45A2F",
                ),
            )
    wb.save(path)


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=_HEADERS)
        writer.writeheader()
        for row in rows:
            writer.writerow({h: row.get(h, "") for h in _HEADERS})


def _slot_fill_keywords_base(vault: Path, csv_path: Path) -> None:
    """Update wiki/meta/keywords.base so its data source points at the CSV.

    Best-effort: if the file doesn't exist or the patch fails, log and continue.
    Obsidian Bases supports both a YAML and JSON-ish format; we do a simple
    line-level substitution for the ``source:`` field so we don't depend on a
    YAML dep.
    """
    base_path = vault / "wiki" / "meta" / "keywords.base"
    if not base_path.exists():
        return
    try:
        text = base_path.read_text(encoding="utf-8")
    except OSError:
        return
    rel_csv = csv_path.relative_to(vault).as_posix() if csv_path.is_absolute() else str(csv_path)
    # Replace any existing source line, or append.
    new_line = f"source: {rel_csv}"
    if re.search(r"^\s*source:\s*.+$", text, flags=re.MULTILINE):
        text = re.sub(r"^\s*source:\s*.+$", new_line, text, count=1, flags=re.MULTILINE)
    else:
        text = text.rstrip() + "\n" + new_line + "\n"
    base_path.write_text(text, encoding="utf-8")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--vault", required=True)
    parser.add_argument("--out-xlsx", default=None)
    parser.add_argument("--out-csv", default=None)
    args = parser.parse_args(argv)

    vault = Path(args.vault).expanduser().resolve()
    raw_dir = vault / ".raw" / "sources" / "dataforseo"
    today = date.today().isoformat()

    competitor_files = _load_competitor_files(raw_dir)
    if not competitor_files:
        print(f"ERROR: no competitor-kw-*.json files in {raw_dir}. Run pull_competitor_kw.py first.", file=sys.stderr)
        return 1

    site_file = _load_site_file(raw_dir)
    site_root: str | None = None
    site_rows: list[dict[str, Any]] = []
    if site_file:
        site_data = json.loads(site_file.read_text(encoding="utf-8"))
        site_root = site_data.get("domain")
        for item in _iter_items(site_data):
            row = _extract_row(site_root or "self", item)
            if row:
                site_rows.append(row)

    competitor_rows: list[dict[str, Any]] = []
    for f in competitor_files:
        data = json.loads(f.read_text(encoding="utf-8"))
        domain = data.get("domain") or f.stem
        for item in _iter_items(data):
            row = _extract_row(domain, item)
            if row:
                competitor_rows.append(row)

    if not competitor_rows:
        print("ERROR: no usable rows extracted from competitor files. Are the JSONs empty?", file=sys.stderr)
        return 1

    aggregated = _aggregate(site_root, site_rows, competitor_rows)
    sheets = _categorize(aggregated)

    out_xlsx = Path(args.out_xlsx).expanduser().resolve() if args.out_xlsx else (vault / f"keywords-{today}.xlsx")
    out_csv = Path(args.out_csv).expanduser().resolve() if args.out_csv else (vault / f"keywords-{today}.csv")
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    out_csv.parent.mkdir(parents=True, exist_ok=True)

    _write_xlsx(out_xlsx, sheets)
    _write_csv(out_csv, sheets["All Keywords"])
    _slot_fill_keywords_base(vault, out_csv)

    # Console summary.
    print(f"\nWrote {out_xlsx}")
    print(f"Wrote {out_csv}")
    print("\nSheet row counts:")
    for name, rows in sheets.items():
        print(f"  {name:<18} {len(rows):>6}")
    print("\nTop 5 High Opportunity:")
    for r in sheets["High Opportunity"][:5]:
        kw = r["keyword"][:50]
        print(f"  - {kw:<52} vol={r['search_volume']:<6} ours={r['our_position']!s:<5} best_comp={r['best_competitor_position']:<3} score={r['opportunity_score']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
