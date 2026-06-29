#!/usr/bin/env python3
"""
Step 4 — mine People-Also-Ask + related searches for the top-volume keywords.

Reads the top-N highest-volume keywords from the dedup'd XLSX (or its CSV
mirror) and queries Google SERP via DataForSEO for each. Collects:

- ``people_also_ask`` SERP elements (the question + each expanded answer's
  source domain & URL)
- ``related_searches`` SERP elements

Outputs both a structured JSON archive and a markdown digest grouped by topic
with frequency counts. The digest is what the ``beast-planner`` subagent reads
in Step 6 to identify content gaps.

Inputs
------
- ``--vault``: vault root.
- ``--top-n``: number of seed keywords to mine (default 100).
- ``--location`` / ``--language``: standard DataForSEO codes.
- ``--cost-cap`` / ``--total-cap``: USD caps.

Outputs
-------
- ``<vault>/.raw/sources/dataforseo/paa-<date>.json``      (per-keyword PAA + related)
- ``<vault>/.raw/sources/dataforseo/paa-digest-<date>.md`` (markdown digest)

Cost
----
SERP regular ~$0.0006/call. 100 keywords ~ $0.06.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path
from typing import Any, Iterable

import _dfs_client as dfs

ENDPOINT = "/v3/serp/google/organic/live/regular"


def _slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")[:60] or "x"


def _read_top_keywords_from_xlsx(xlsx_path: Path, top_n: int) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if "All Keywords" not in wb.sheetnames:
        return []
    ws = wb["All Keywords"]
    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        return []
    headers = [str(h) if h is not None else "" for h in headers]
    out: list[dict[str, Any]] = []
    for row in rows_iter:
        record = dict(zip(headers, row))
        out.append(record)
    out.sort(key=lambda r: -(int(r.get("search_volume") or 0)))
    return out[:top_n]


def _read_top_keywords_from_csv(csv_path: Path, top_n: int) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with csv_path.open("r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                row["search_volume"] = int(row.get("search_volume") or 0)
            except (TypeError, ValueError):
                row["search_volume"] = 0
            rows.append(row)
    rows.sort(key=lambda r: -r["search_volume"])
    return rows[:top_n]


def _walk_serp_for_paa_and_related(items: Iterable[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[str]]:
    """Walk SERP items (which can nest) and return (paa_questions, related_searches)."""
    paa: list[dict[str, Any]] = []
    related: list[str] = []
    stack: list[dict[str, Any]] = list(items)
    while stack:
        item = stack.pop()
        if not isinstance(item, dict):
            continue
        t = item.get("type")
        if t == "people_also_ask":
            for el in item.get("items") or []:
                if isinstance(el, dict) and el.get("type") == "people_also_ask_element":
                    question = el.get("title")
                    expanded = el.get("expanded_element") or []
                    sources = []
                    for sub in expanded:
                        if isinstance(sub, dict):
                            sources.append({
                                "url": sub.get("url"),
                                "domain": sub.get("domain"),
                                "title": sub.get("title"),
                            })
                    if question:
                        paa.append({"question": question, "sources": sources})
        elif t == "related_searches":
            for q in item.get("items") or []:
                if isinstance(q, str) and q.strip():
                    related.append(q.strip())
        # Some SERP item types nest further (e.g. inside ``items``)
        nested = item.get("items")
        if isinstance(nested, list):
            stack.extend(x for x in nested if isinstance(x, dict))
    return paa, related


def _topic_bucket(question: str) -> str:
    q = question.lower()
    if any(w in q for w in ["how", "guide", "tutorial", "step"]):
        return "How / Guide"
    if any(w in q for w in ["best", "top", "vs", "versus", "review"]):
        return "Comparison / Review"
    if any(w in q for w in ["what", "definition", "mean", "is a", "is an"]):
        return "Definition / What-is"
    if any(w in q for w in ["why"]):
        return "Why / Reasoning"
    if any(w in q for w in ["where", "near", "location"]):
        return "Where / Location"
    if any(w in q for w in ["when", "time", "season"]):
        return "When / Timing"
    if any(w in q for w in ["cost", "price", "how much", "fee"]):
        return "Pricing"
    return "Other"


def _write_digest(path: Path, paa_data: list[dict[str, Any]], related_data: list[dict[str, Any]]) -> None:
    # PAA frequency
    question_counter: Counter[str] = Counter()
    question_to_keywords: dict[str, set[str]] = defaultdict(set)
    bucket_to_questions: dict[str, list[tuple[str, int]]] = defaultdict(list)
    for record in paa_data:
        seed = record["seed_keyword"]
        for q in record.get("questions") or []:
            question_counter[q["question"]] += 1
            question_to_keywords[q["question"]].add(seed)
    for question, count in question_counter.most_common():
        bucket_to_questions[_topic_bucket(question)].append((question, count))

    related_counter: Counter[str] = Counter()
    related_to_keywords: dict[str, set[str]] = defaultdict(set)
    for record in related_data:
        seed = record["seed_keyword"]
        for q in record.get("related") or []:
            related_counter[q] += 1
            related_to_keywords[q].add(seed)

    lines: list[str] = []
    lines.append("---")
    lines.append("brain_schema: marketing-brain.v1")
    lines.append("type: source")
    lines.append('title: "PAA Mining Digest"')
    lines.append(f"created: {date.today().isoformat()}")
    lines.append(f"updated: {date.today().isoformat()}")
    lines.append("tags:")
    lines.append("  - source/dataforseo")
    lines.append("  - paa")
    lines.append("status: developing")
    lines.append("---")
    lines.append("")
    lines.append("# PAA Mining Digest")
    lines.append("")
    lines.append(f"Mined {len(paa_data)} top-volume keywords. {len(question_counter)} unique PAA questions, {len(related_counter)} unique related searches.")
    lines.append("")
    lines.append("## People Also Ask — by topic")
    lines.append("")
    for bucket, items in bucket_to_questions.items():
        lines.append(f"### {bucket}")
        lines.append("")
        for question, count in items[:50]:
            seeds = sorted(question_to_keywords[question])
            seed_preview = ", ".join(seeds[:3]) + (f" +{len(seeds)-3}" if len(seeds) > 3 else "")
            lines.append(f"- ({count}x) **{question}** _seeds: {seed_preview}_")
        lines.append("")

    lines.append("## Related Searches — top 100")
    lines.append("")
    for q, count in related_counter.most_common(100):
        seeds = sorted(related_to_keywords[q])
        seed_preview = ", ".join(seeds[:3]) + (f" +{len(seeds)-3}" if len(seeds) > 3 else "")
        lines.append(f"- ({count}x) {q} _seeds: {seed_preview}_")

    dfs.write_private_text(path, "\n".join(lines) + "\n")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--vault", required=True)
    parser.add_argument("--top-n", type=int, default=100)
    parser.add_argument("--location", type=int, default=2124)
    parser.add_argument("--language", default="en")
    parser.add_argument("--cost-cap", type=float, default=0.50)
    parser.add_argument("--total-cap", type=float, default=5.00)
    parser.add_argument("--xlsx", default=None, help="Explicit path to keywords XLSX (defaults to most recent)")
    parser.add_argument("--csv", default=None, help="Explicit path to keywords CSV (fallback if no XLSX)")
    parser.add_argument("--dry-run", action="store_true", help="Preview SERP/PAA calls and estimated cost without calling DataForSEO")
    args = parser.parse_args(argv)

    dfs.set_caps(per_call=args.cost_cap, total=args.total_cap)

    vault = Path(args.vault).expanduser().resolve()
    raw_dir = vault / ".raw" / "sources" / "dataforseo"
    today = date.today().isoformat()

    # Locate inputs.
    xlsx_path = Path(args.xlsx).expanduser().resolve() if args.xlsx else None
    if xlsx_path is None:
        candidates = sorted(vault.glob("keywords-*.xlsx"))
        xlsx_path = candidates[-1] if candidates else None
    csv_path = Path(args.csv).expanduser().resolve() if args.csv else None
    if csv_path is None:
        candidates = sorted(vault.glob("keywords-*.csv"))
        csv_path = candidates[-1] if candidates else None

    keywords: list[dict[str, Any]] = []
    if xlsx_path and xlsx_path.exists():
        keywords = _read_top_keywords_from_xlsx(xlsx_path, args.top_n)
    if not keywords and csv_path and csv_path.exists():
        keywords = _read_top_keywords_from_csv(csv_path, args.top_n)

    if not keywords:
        print(f"ERROR: no keywords found in {xlsx_path} / {csv_path}. Run build_keyword_xlsx.py first.", file=sys.stderr)
        return 1

    if args.dry_run:
        estimated = len(keywords) * 0.0006
        print("DRY RUN: no DataForSEO calls made")
        print(f"Endpoint: {ENDPOINT}")
        print(f"Keyword seeds: {len(keywords)}")
        print(f"Estimated spend: ${estimated:.4f} before DataForSEO account pricing adjustments")
        for row in keywords[:10]:
            print(f"- {row.get('keyword')}")
        return 0

    dfs.require_credentials()

    paa_records: list[dict[str, Any]] = []
    related_records: list[dict[str, Any]] = []
    out_dir = raw_dir
    out_dir.mkdir(parents=True, exist_ok=True)

    try:
        for kw_record in keywords:
            seed = kw_record.get("keyword") or ""
            if not seed:
                continue
            payload = [{
                "keyword": seed,
                "location_code": args.location,
                "language_code": args.language,
                "depth": 10,
            }]
            save_to = out_dir / f"paa-serp-{_slug(seed)}-{today}.json"
            try:
                data = dfs.call(ENDPOINT, payload, label=f"paa/{seed}", save_to=save_to)
            except dfs.DataForSEOError as exc:
                print(f"WARN: skipping '{seed}': {exc}", file=sys.stderr)
                continue
            tasks = data.get("tasks") or []
            if not tasks:
                continue
            results = tasks[0].get("result") or []
            if not results:
                continue
            items = results[0].get("items") or []
            paa, related = _walk_serp_for_paa_and_related(items)
            paa_records.append({"seed_keyword": seed, "questions": paa})
            related_records.append({"seed_keyword": seed, "related": related})
    except dfs.CostCapExceeded as exc:
        print(f"ERROR: {exc}. Writing partial output.", file=sys.stderr)
        _write_outputs(out_dir, today, paa_records, related_records, partial=True)
        return 1

    _write_outputs(out_dir, today, paa_records, related_records, partial=False)
    print(f"\nMined {len(paa_records)} keywords. Cost ${dfs.total_cost():.4f}.")
    print(f"Wrote {out_dir / f'paa-{today}.json'}")
    print(f"Wrote {out_dir / f'paa-digest-{today}.md'}")
    return 0


def _write_outputs(
    out_dir: Path,
    today: str,
    paa_records: list[dict[str, Any]],
    related_records: list[dict[str, Any]],
    *,
    partial: bool,
) -> None:
    archive = {
        "generated_at": today,
        "partial": partial,
        "paa": paa_records,
        "related_searches": related_records,
    }
    dfs.write_private_text(out_dir / f"paa-{today}.json", json.dumps(archive, indent=2) + "\n")
    _write_digest(out_dir / f"paa-digest-{today}.md", paa_records, related_records)


if __name__ == "__main__":
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    raise SystemExit(main())
